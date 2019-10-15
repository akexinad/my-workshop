import openpyxl
import os

# https://www.youtube.com/watch?v=q6Mc_sAPZ2Y

# print(openpyxl.__version__)

filename = 'pyxl.xlsx'
src_path = "C:\\Users\\cecd304\\me\\code\\workshop\\spreadsheet_code\\"
os.chdir(src_path)

wb = openpyxl.load_workbook(filename)

# print(type(wb))

sheet = wb['Sheet1']

sheet['d4'].value = 'HELLO WORLD'

# wb.save(filename)

# print(sheet1['b1:b7'])

# print(sheet.title)


# wb.save(filename)

# print(type(sheet['a1'].value))


# helper function to return 
def cell(col, row):
    return col + str(row)


# range hs to be from 1 to 5 since the excel rows start from 1 and not 0;
for i in range(1, 5):
    res = cell('a', i)
    print(sheet[res].value)

print(sheet.max_row)
print(sheet.max_column)